#!/usr/bin/env python
# _*_ coding:utf-8 _*_
__author__ = 'Sugar'

import os
import xlrd
import shutil
from config import setting
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import RED, GREEN, DARKYELLOW

from testframe.common.config_util import ConfigUtil


# --------- 读取config.ini配置文件 ---------------
ConfigUtil.get_ini("tester", "name")


class WriteExcel:
    """文件写入数据"""

    def __init__(self, fileName):
        self.filename = fileName
        if not os.path.exists(self.filename):
            # 文件不存在，则拷贝模板文件至指定报告目录下
            shutil.copyfile(setting.SOURCE_FILE, setting.TARGET_FILE)
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active

    def write_data(self, row_n, value):
        """
        写入测试结果
        :param row_n:数据所在行数
        :param value: 测试结果值
        :return: 无
        """
        font_GREEN = Font(name='宋体', color=GREEN, bold=True)
        font_RED = Font(name='宋体', color=RED, bold=True)
        font1 = Font(name='宋体', color=DARKYELLOW, bold=True)
        align = Alignment(horizontal='center', vertical='center')
        # 获数所在行数
        L_n = "L" + str(row_n)
        M_n = "M" + str(row_n)
        if value == "PASS":
            self.ws.cell(row_n, 12, value)
            self.ws[L_n].font = font_GREEN
        if value == "FAIL":
            self.ws.cell(row_n, 12, value)
            self.ws[L_n].font = font_RED
        self.ws.cell(row_n, 13, name)
        self.ws[L_n].alignment = align
        self.ws[M_n].font = font1
        self.ws[M_n].alignment = align
        self.wb.save(self.filename)


class ReadExcel:
    """读取excel文件数据"""

    def __init__(self, fileName, SheetName="Sheet1"):
        self.data = xlrd.open_workbook(fileName)
        self.table = self.data.sheet_by_name(SheetName)

        # 获取总行数、总列数
        self.nrows = self.table.nrows
        self.ncols = self.table.ncols

    def read_data(self):
        if self.nrows > 1:
            # 获取第一行的内容，列表格式
            keys = self.table.row_values(0)
            listApiData = []
            # 获取每一行的内容，列表格式
            for col in range(1, self.nrows):
                values = self.table.row_values(col)
                # keys，values组合转换为字典
                api_dict = dict(zip(keys, values))
                listApiData.append(api_dict)
            return listApiData
        else:
            print("表格是空数据!")
            return None
